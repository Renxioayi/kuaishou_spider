import requests
from lxml import html
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib import request
import time
import re
def set_columns_width(sheet, columns_width):
    # 定义列的长度
    for index, width in enumerate(columns_width):
        sheet.column_dimensions[get_column_letter(index + 1)].width = width


def getInitInfo():
    videos = {}
    try:
        workbook = load_workbook('视频统计.xlsx')
        sheet = workbook.active
        columns = ['url']
        for row_index in range(1, sheet.max_row + 1):
            video = {}
            for index, column in enumerate(columns):
                video[column] = sheet.cell(
                    row=row_index, column=index + 1).value
            # print(video['url'])
            video['url'] = video['url'].strip()
            video['watch'] = ''
            video['like'] = ''
            videos[video['url']] = video
    except FileNotFoundError:
        print('视频统计.xlsx文件不存在,请建立文件或重命名xlsx文件')

    return videos

def getCollectInfo():
    videos = getInitInfo()
    for i in videos:
        url_length = len(i)
        _watch = '0'
        _like = 'link_invalid'

        if url_length < 70:
            # 爬取短链接
            res = requests.get(videos[i]['url']).content
            sel = html.fromstring(res)
            if (sel.xpath('//span[@class="watching"]/text()')):
                i_watch = (sel.xpath('//span[@class="watching"]/text()')[0]).strip()
                _watch = re.findall('[0-9\.a-z]+', i_watch)[0]
            if (sel.xpath('//span[@class="like-count"]/text()')):
                _like = sel.xpath('//span[@class="like-count"]/text()')[0]
            print('已爬取短链{}'.format(i))
            # a = 1


        elif url_length>=70:
                res = request.urlopen(videos[i]['url']).read()
                sel = html.fromstring(res)
                time.sleep(5)
                a = "test"
                if (sel.xpath('//a[@class="room-link"]/text()')):
                    tmp = sel.xpath('//a[@class="room-link"]/text()')[0]
                    a = tmp.strip()

                if a == "前往她的个人中心" or a =="前往他的个人中心":
                    # 爬取长链电脑版
                    if (sel.xpath('//span[@class="watching"]/text()')):
                        i_watch = (sel.xpath('//span[@class="watching"]/text()')[0]).strip()
                        _watch = re.findall('[0-9\.a-z]+',i_watch)[0]
                    if (sel.xpath('//span[@class="like-count"]/text()')):
                        _like = sel.xpath('//span[@class="like-count"]/text()')[0]
                    print('已爬取长链电脑版{}'.format(i))

                else:
                    # 爬取长链手机版
                    if (sel.xpath('//div[@class="comments-num"]/span[3]/text()')):
                        i_watch = (sel.xpath('//div[@class="comments-num"]/span[3]/text()')[0]).strip()
                        _watch = re.findall('[0-9\.a-z]+',i_watch)[0]
                    if (sel.xpath('//div[@class="comments-num"]/span[2]/text()')):
                        _like = sel.xpath('//div[@class="comments-num"]/span[2]/text()')[0]
                    print('已爬取长链手机版{}'.format(i))

        videos[i]['watch'] = _watch.replace(" ", "").replace("\n", "")
        videos[i]['like'] = _like.replace(" ", "")

        # print('已爬取{}'.format(i))

    return videos


def generate():
    workbook = Workbook()
    sheet = workbook.active
    set_columns_width(sheet, [50, 10, 15])
    sheet.append(['链接', '观看数', '赞'])

    videos = getCollectInfo()
    for i in videos:
        sheet.append([videos[i]['url'], videos[i]['watch'], videos[i]['like']])

    try:
        workbook.save('output.xlsx')
        workbook.close()
    except PermissionError:
        print('无法写入')


if __name__ == '__main__':
    print("======================欢迎使用南讯快手视频播放量爬虫======================")
    generate()
    print("======================爬取完成======================")